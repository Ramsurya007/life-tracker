from openpyxl import Workbook
from django.utils import timezone
from calendar import monthrange

from core.services.scores import calculate_daily_score
from core.services.summaries import weekly_summary, monthly_summary
from core.models import NoFapEntry, DailyNote


def export_life_excel(user):
    wb = Workbook()

    # ---------------- Sheet 1: Daily Log ----------------
    ws_daily = wb.active
    ws_daily.title = "Daily Log"

    ws_daily.append(["Date", "Score %", "Status", "NoFap", "Note"])

    today = timezone.now().date()
    year = today.year

    for month in range(1, 13):
        days = monthrange(year, month)[1]
        for day in range(1, days + 1):
            date = timezone.datetime(year, month, day).date()
            score = calculate_daily_score(user, date)

            nofap = NoFapEntry.objects.filter(user=user, date=date, is_clean=True).exists()
            note = DailyNote.objects.filter(user=user, date=date).first()

            ws_daily.append([
                str(date),
                score["score"],
                score["status"],
                "Clean" if nofap else "Relapse",
                note.note if note else ""
            ])

    # ---------------- Sheet 2: Weekly ----------------
    ws_weekly = wb.create_sheet("Weekly Summary")
    ws_weekly.append(["Date", "Score %", "Status"])

    for day in weekly_summary(user):
        ws_weekly.append([
            str(day["date"]),
            day["score"],
            day["status"]
        ])

    # ---------------- Sheet 3: Monthly ----------------
    ws_monthly = wb.create_sheet("Monthly Summary")
    monthly = monthly_summary(user, today.year, today.month)

    ws_monthly.append(["Metric", "Value"])
    ws_monthly.append(["Average Score", monthly["avg_score"]])
    ws_monthly.append(["Best Day Score", monthly["best_day"]])
    ws_monthly.append(["Worst Day Score", monthly["worst_day"]])
    ws_monthly.append(["NoFap Clean Days", monthly["nofap_clean_days"]])
    ws_monthly.append(["Total Days", monthly["total_days"]])

    return wb
